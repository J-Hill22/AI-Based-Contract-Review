import openpyxl
import nltk
import re
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords
from gensim.models import KeyedVectors
import numpy as np
from sklearn.neighbors import NearestNeighbors

def extract_sheet_data(sheet):

    data = {
        "Auburn's Preferred Language": [],
        'Common Problems' : [],
        'Why': [],
        '1st response to Sponsor' : []
    }
    # Quick fix for ensuring both cell title forms are adressed, 
    # could/should probably be improved in future
    auburn_title = ["Auburn's Preferred Language", 'Auburn Preferred Language']

    current_key = None
    for row in sheet.iter_rows(values_only = True):
        if not row[1]: continue #skip empty rows in second columnn

        if row[1] in auburn_title: 
            current_key =  "Auburn's Preferred Language"
            continue

        if row[1] in data.keys():
            current_key = row[1]
            continue

        if current_key == 'Common Problems':
            data['Common Problems'].append(row[1])
            data['Why'].append(row[2] if len(row) > 2 and row[2] else None)
            data['1st response to Sponsor'].append(row[3] if len(row) > 3 and row[3] else None)
        elif current_key == "Auburn's Preferred Language":
            data[current_key].append(row[1])

    return data

# This function exists solely for testing purposes
# used to generate a readable representation of the parsed T&Cs matrix
def pretty_print_nested_values(dictionary):
    output = []
    for page_title, inner_dict in dictionary.items():
        output.append(page_title)
        for cell_title, values in inner_dict.items():
            output.append("\t" + cell_title + ":")
            for value in values:
                str_value = str(value) if value is not None else "None"
                output.append("\t\t" + str_value)
        output.append('')  # Add a blank line for separation
    return "\n".join(output)

# Function to get vector representation for a sentence
def get_sentence_vector(sentence, model):
    words = [word for word in nltk.word_tokenize(sentence) if word in model.key_to_index and word not in stopwords.words('english')]
    if not words:
        return np.zeros(model.vector_size)  # return a zero vector if no word is found
    vectors = [model[word] for word in words]
    return sum(vectors) / len(vectors)

def is_sentence_problematic(sentence, model, nn_model, threshold):
    """Returns True if the sentence is close enough to a known "Common Problem" sentence.
    The threshold can be adjusted based on requirements."""
    sentence_vector = get_sentence_vector(sentence, model)
    distance, _ = nn_model.kneighbors([sentence_vector])
    return distance[0][0] < threshold

# This function generates a similarity score when comparing two sentences of a contract
def jaccard_similarity(sentence1, sentence2):
    # Tokenize sentences
    words1 = set(word_tokenize(sentence1.lower()))
    words2 = set(word_tokenize(sentence2.lower()))

    # Remove stopwords
    stop_words =  set(stopwords.words('english'))
    words1 = words1- stop_words
    words2 = words2- stop_words

    # Calculate Jaccard Similarity
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    return len(intersection) / len(union)

# This function flags sentences based on how 'similar' they are to language 
# in the common problems cell of each T&Cs sheet
def flag_sentences(tnc_dictionary):

    with open('contract_to_txt.txt', 'r', encoding = 'utf-8') as file:
        contract_text = file.read()

    contract_sentences = sent_tokenize(contract_text)

    threshold = 0.20
    flagged_sentences = {}

    model = KeyedVectors.load_word2vec_format("G:\Downloads\\archive\GoogleNews-vectors-negative300.bin", binary=True)

    # Nearest Neighbor model setup
    nn_model = NearestNeighbors(n_neighbors=1, metric='cosine', algorithm='brute')

    common_problems_sentences = []
    for key in tnc_dictionary:
        common_problems_sentences.extend(tnc_dictionary[key]['Common Problems'])
    #print(common_problems_sentences)
    common_problems_vectors = [get_sentence_vector(sentence, model) for sentence in common_problems_sentences]
    #print(common_problems_vectors[:10])
    nn_model.fit(common_problems_vectors)

    for key, sub_dict in tnc_dictionary.items():
        if "Common Problems" in sub_dict:
            sheet_name = key
            preferred_language = sub_dict.get("Auburn's Preferred Language", "N/A")
            why_list = sub_dict.get("Why", [])
            response_list = sub_dict.get("1st response to Sponsor", [])
            
            for sentence in contract_sentences:
                for i, value in enumerate(sub_dict["Common Problems"]):
                    if jaccard_similarity(sentence, value) > 0.12:
                        if sentence not in flagged_sentences:
                            flagged_sentences[sentence] = []
                        flagged_sentences[sentence] = {
                            "Problem Category" : sheet_name,
                            "Common Problems" : value,
                            "Preferred Language" : preferred_language,
                            "Why" : why_list[i] if i < len(why_list) else "N/A",
                            "1st response to Sponsor" : response_list[i] if i < len(response_list) else "N/A"
                        }

    return flagged_sentences

# Compares the entire contract to 'flagged_sentences' and inserts the relevant data
# into the txt file below the matched flagged sentence
def _flag_problem_language(tnc_path_in):
    wb = openpyxl.load_workbook(tnc_path_in)

    sheet_names = wb.sheetnames

    tnc_dictionary = {}

    for sheet_name in sheet_names:
        if sheet_name != "INDEX" and sheet_name != "template" and sheet_name != "CONTACTS":
            sheet = wb[sheet_name]
            sheet_data = extract_sheet_data(sheet)
            tnc_dictionary[sheet_name] = sheet_data

    flagged_sentences = flag_sentences(tnc_dictionary)
    with open('contract_to_txt.txt', 'r', encoding = 'utf-8') as file:
        lines = file.read()

    token_lines = sent_tokenize(lines)
    modded_lines = []

    for tokenized_line in token_lines:
        matched = False
        for key in flagged_sentences:
            if key.strip() == tokenized_line.strip():
                matched = True
                modded_lines.append("\t\t[POTENTIAL PROBLEMATIC LANGUAGE DETECTED]")
                modded_lines.append(tokenized_line)
                for k, v in flagged_sentences[key].items():
                    modded_lines.append(f"\t\t{k}: {v}")
                break
        if not matched:
            modded_lines.append(tokenized_line)

    with open('flagged_contract_to_txt.txt', 'w', encoding = 'utf-8') as f:
        for eachline in modded_lines:
            f.write(eachline + '\n')